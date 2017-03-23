# encoding: utf-8

from openpyxl import Workbook

# 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
wb = Workbook()

ws = wb.get_active_sheet()
print(ws.title)
#ws.title = 'New Title'  # 设置worksheet的标题

# 设置单元格的值
#ws.cell('D3').value = 4
#ws.cell(row=3, column=1).value = 6

new_ws = wb.create_sheet(title='new_sheet')
for row in range(1, 10):
    for col in range(1, 10):
        new_ws.cell(row = row, column = col).value = row * col

new_ws1=wb.create_sheet(title='new_sheet1')
for row in range(1,10):
    for col in range(1,row+1):
        new_ws1.cell(row = row,column = col).value = row * col

# 最后一定要保存！
wb.save(filename='write.xlsx')